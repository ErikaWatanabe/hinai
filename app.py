import glob, shutil, os
import pandas as pd
from datetime import datetime
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font


# 年、月、日を取得し作成するシフトが上旬か下旬の判定
period = "下旬" #下旬
now = datetime.now()
current_year = now.year
current_month = now.month
current_day = now.day
if(current_day > 18): #18日以降なら来月のシフト
    period = "上旬" #上旬
    if(current_month == 12): #12月なら年越し
        current_year += 1
        current_month = 1
    else:
        current_month += 1


# 指定したフォルダのパス
read_excel_path = 'read_excel' # 希望シフトが書いてあるファイルが入ってるフォルダのパス 
format_excel_path = 'format_excel' # シフト表のテンプレが入ってるフォルダのパス
write_file_path = f'{current_month}月{period}のシフト.xlsx'  # 新しいファイルのパス

# フォルダ内のExcelファイル（拡張子が .xlsx のもの）を検索
read_excel_file = glob.glob(f'{read_excel_path}/*.xlsx')
format_excel_file = glob.glob(f'{format_excel_path}/*.xlsx')

shutil.copy(format_excel_file[0], write_file_path)

def find_file(excelfile):
    if excelfile:
        file_path = excelfile[0]
        df = pd.read_excel(file_path)
        print(f'読み込んだファイル: {file_path}')
    else:
        print('Excelファイルが見つかりませんでした。')
    return df

df_read = find_file(read_excel_file)
df_write = pd.read_excel(write_file_path)

# 各行をリストとして格納
rows_as_list = []
name_list = []
for index, row in df_read.iterrows():
    rows_as_list.append(row.tolist())  # 行をリストに変換して追加

# # name_listに名前を格納
# for row in rows_as_list:
#     name_list.append(row[2])
# print(rows_as_list[1])

# シフト表書き込み開始
# 年月書き込み
workbook = openpyxl.load_workbook(write_file_path)
sheet = workbook.active
sheet['B1'] = current_year
sheet['F1'] = current_month

# 日付書き込み
num_of_cell = 15
start_column = 2
start_row = 3
day_list = []

if period == "下旬":
    day_list = [16, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27, 28, 29, 30]
    if current_month in (1, 3, 5, 7, 8, 10, 12): #31日までの月は31日をday_listに足す
        num_of_cell = 16
        day_list.append(31)
else:
    day_list = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15]

def convert_font(color_code, c_date, c_day_of_week): # 土日のフォント色を変える関数
    c_date.font = Font(name="HG丸ｺﾞｼｯｸM-PRO", size=10, bold=True, color=color_code)
    c_day_of_week.font = Font(name="HG丸ｺﾞｼｯｸM-PRO", size=10, bold=True, color=color_code)

for i in range(num_of_cell):
    # 日付
    cell_date = sheet.cell(row=start_row, column=start_column + 2*i)
    cell_date.value = day_list[i]

    # 曜日
    cell_day_of_week = sheet.cell(row=start_row, column=start_column + 2*i + 1)
    date = datetime(current_year, current_month, day_list[0]+i)
    weekdays_jp = ["月", "火", "水", "木", "金", "土", "日"]
    cell_day_of_week.value = weekdays_jp[date.weekday()]
    # 土曜と日曜の色指定
    if date.weekday() == 5:  # 土曜日
        convert_font("4BACC6", cell_date, cell_day_of_week)
    elif date.weekday() == 6:  # 日曜日
        convert_font("FF0000", cell_date, cell_day_of_week)

def split_time(num, data_row, start_row, _sheet, period):
    if data_row[3] == period:
        for i in range(num):
            start = 0
            end = 0
            time_range = data_row[4+i]
            if not pd.isna(time_range):
                start_time, end_time = time_range.split("~")
                start_hour, start_minute = start_time.split(":")
                end_hour, end_minute = end_time.split(":")
                start = int(start_hour) if start_minute == "00" else int(start_hour) + 0.5
                end = int(end_hour) if end_minute == "00" else int(end_hour) + 0.5
                # print(f"start:{start}({type(start)}), end:{end}({type(end)})")
                cell_start = _sheet.cell(row=start_row, column=2 + 2*i)
                cell_start.value = start
                cell_end = _sheet.cell(row=start_row, column=2 + 2*i + 1)
                cell_end.value = end
    else: #上旬下旬を間違って選択していた場合
        print(f"{row[2]}は上旬下旬の選択が間違ってます")
        cell = _sheet.cell(row=start_row, column=34)
        cell.value = f"{row[3]}を選択しています"
        cell.font = Font(color="FF0000")
        
# それぞれの希望シフト書き込み
# writeファイルの1列目に一致する名前があるか検索
for row in rows_as_list:
    count = 0
    find = False
    for write_sel_1 in df_write.iloc[:, 0]:  # 1列目を取得
        count += 1
        if row[2] == write_sel_1:  # 一致する名前がname_listにあるか確認
            print(f"{row[2]}はwriteファイルの{count+1}行目にありました")
            find = True
            split_time(num_of_cell, row, count+1, sheet, period) # 時間分割して書き込み
            break
    if not (find):
        print(f"{row[2]}を見つけられませんでした")



# 変更を保存
workbook.save(write_file_path)
